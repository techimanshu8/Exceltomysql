from openpyxl import load_workbook
import mysql.connector
#put name of file instead of example.xlsx
wb= load_workbook("example.xlsx") 
ws=wb.active
def connecttodb(list):
    #put your database name in place of data1
    mydb = mysql.connector.connect(host="localhost", user ="root", passwd="" ,database="data1")
    #uncomment print() if you want to see which raw is going to be saved
    #print(list)
    mycurser = mydb.cursor()
    # put your Table name  and  fields name which is exectly in your structure database
    #you can add more fields and values as per your requirements
    #this program puts list[0](column 'A') to field1 and similarly for list[1]='B' and so on
    mycurser.execute(f"INSERT INTO tablename (field1, field2, field3, field4) VALUES ('{list[0]}', '{list[1]}', '{list[2]}', '{list[3]}')")
    mydb.commit()
    mycurser.close()
    mydb.close()
#place your desired value of Fromraw to Toraw Which you want to add in database
#remember if you want to save from raw 1 to 5 then you must write range from 1 to (5+1=6)

for raw in range(Fromraw=1,Toraw=6):
    listcoln=[]
    #put column names insted of 'A' and 'E' as per your Requirements
    #to variable valid only upto 'Z'
    from='A'
    to='E'
    for coln in range(ord(from),ord(to)):
        listcoln.append(ws[chr(coln)+str(raw)].value)
    connecttodb(listcoln)
    
#put name of file instead of example.xlsx
wb.save("example.xlsx")
